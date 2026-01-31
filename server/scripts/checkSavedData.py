import openpyxl
import os

file_path = r'c:\Users\aniam\Desktop\PROJET TECHWEB\server\data\BDD1.xlsx'

if not os.path.exists(file_path):
    print("File not found at:", file_path)
else:
    print("File found! Loading...")
    wb = openpyxl.load_workbook(file_path)
    ws = wb['Utisers']
    
    # Find user with ID 1769718198386
    user_id = '1769718198386'
    found = False
    
    for row in ws.iter_rows(min_row=2, values_only=False):
        if str(row[0].value) == user_id:
            found = True
            print(f"\n✓ User found! Here's the saved data:\n")
            print(f"ID: {row[0].value}")
            print(f"Prénom: {row[1].value}")
            print(f"Nom: {row[2].value}")
            print(f"Email: {row[3].value}")
            print(f"Mot de passe (hash): {row[4].value}")
            print(f"Téléphone (numTel): {row[5].value}")
            print(f"Langue: {row[6].value}")
            print(f"Thème: {row[7].value}")
            print(f"Date création: {row[8].value}")
            print(f"Rôle: {row[9].value}")
            print(f"Adresse: {row[10].value}")
            print(f"Ville: {row[11].value}")
            print(f"Code postal: {row[12].value}")
            print(f"Pays: {row[13].value}")
            print(f"Nom carte: {row[14].value}")
            print(f"Numéro carte: {row[15].value}")
            print(f"Date expiration: {row[16].value}")
            print(f"CVV: {row[17].value}")
            break
    
    if not found:
        print(f"\n✗ User with ID {user_id} not found in database")
        print("\nFirst 5 users in database:")
        count = 0
        for row in ws.iter_rows(min_row=2, values_only=False):
            if count >= 5:
                break
            print(f"ID: {row[0].value}, Email: {row[3].value}")
            count += 1
